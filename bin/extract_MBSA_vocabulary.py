import pandas as pd
import sys
import os
import re
from openpyxl import load_workbook

MY_PREFIX_LABEL = "mbsa"
MY_NAMESPACE = "http://fd.cvut.cz/chopamax/mbsa/aida"
MY_PREFIX_DECL = f"@prefix {MY_PREFIX_LABEL}: <{MY_NAMESPACE}> .\n"
RDF_PREFIX = "@prefix rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#> .\n"
OWL_PREFIX = "@prefix owl: <http://www.w3.org/2002/07/owl#> .\n"
RDFS_PREFIX = "@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .\n\n"

ONTOLOGY_HEADER = (
    f"{MY_PREFIX_LABEL}: rdf:type owl:Ontology ;\n"
    f"    owl:versionIRI {MY_PREFIX_LABEL}:v0.1 .\n\n"
)

def make_uri_safe(text):
    """
    Convert a string into a URI-safe identifier by replacing whitespace
    with underscores and removing non-alphanumeric characters.
    """
    text = re.sub(r'\s+', '_', text)
    text = re.sub(r'[^a-zA-Z0-9_]', '', text)
    return text

def main():
    if len(sys.argv) != 3:
        print("Usage: python extract_MBSA_vocabulary.py input_file.xlsx output_file.ttl")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    workbook = load_workbook(input_file, data_only=True)
    sheet = workbook.active
    visible_rows = [row[0].row for row in sheet.iter_rows() if not sheet.row_dimensions[row[0].row].hidden]
    print(f"Found {len(visible_rows)} visible rows.", file=sys.stderr)

    df = pd.read_excel(input_file, header=None, names=["Nature", "Parent", "Name"], usecols=[1, 2, 3])
    df = df.iloc[visible_rows]

    print("DataFrame loaded (first 5 visible rows):", file=sys.stderr)
    print(df.head(), file=sys.stderr)

    with open(output_file, "w", encoding="utf-8") as f:

        f.write(MY_PREFIX_DECL)
        f.write(RDF_PREFIX)
        f.write(OWL_PREFIX)
        f.write(RDFS_PREFIX)
        f.write(ONTOLOGY_HEADER)

        for index, row in df.iterrows():
            # Log the raw row (as dictionary)
            print(f"Processing row {index}: {row.to_dict()}", file=sys.stderr)

            # Convert values to strings; if not available, use empty string.
            nature = str(row["Nature"]).strip() if pd.notna(row["Nature"]) else ""
            parent = str(row["Parent"]).strip() if pd.notna(row["Parent"]) else ""
            name = str(row["Name"]).strip() if pd.notna(row["Name"]) else ""
            print(f"Row {index} processed: nature='{nature}', parent='{parent}', name='{name}'", file=sys.stderr)

            if not name:
                print(f"Skipping row {index} because name is empty.", file=sys.stderr)
                continue

            name_uri = make_uri_safe(name)
            parent_uri = make_uri_safe(parent) if parent else ""

            f.write(f"{MY_PREFIX_LABEL}:{name_uri} rdf:type owl:Class ;\n")
            f.write(f"    rdfs:label \"{name}\" ;\n")
            f.write(f"    {MY_PREFIX_LABEL}:nature \"{nature}\" ;\n")
            if nature in ["InConnector", "OutConnector", "AtomicBrickInstance",
                          "CompositeBrickInstance", "Event", "StateVariable", "Value"]:
                f.write(f"    rdfs:subClassOf {MY_PREFIX_LABEL}:{make_uri_safe(nature)} ;\n")

            if parent_uri:
                f.write(f"    {MY_PREFIX_LABEL}:belongsTo {MY_PREFIX_LABEL}:{parent_uri} ;\n")

            f.write("    .\n\n")

    print(f"Vocabulary has been written to {output_file}", file=sys.stderr)

if __name__ == "__main__":
    main()
